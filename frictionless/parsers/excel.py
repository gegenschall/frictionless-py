import os
import sys
import xlrd
import shutil
import atexit
import openpyxl
import datetime
from itertools import chain
from tempfile import NamedTemporaryFile
from ..parser import Parser
from ..system import system
from ..file import File
from .. import exceptions
from .. import dialects
from .. import helpers
from .. import errors


class XlsxParser(Parser):
    Dialect = dialects.ExcelDialect

    # Read

    # TODO: we are not getting stats here
    def read_loader(self):
        source = self.file.source
        loader = system.create_loader(self.file)
        workbook_cache = self.file.dialect.workbook_cache

        # Network
        # Create copy for remote source
        # For remote stream we need local copy (will be deleted on close by Python)
        # https://docs.python.org/3.5/library/tempfile.html#tempfile.TemporaryFile
        if loader.network:
            # Cached
            if workbook_cache is not None and source in workbook_cache:
                loader = system.create_loader(File(source=source, stats=self.file.stats))
            # Not cached
            else:
                prefix = 'tabulator-'
                delete = workbook_cache is None
                loader = system.create_loader(self.file)
                loader.open(mode='b')
                target = NamedTemporaryFile(prefix=prefix, delete=delete)
                shutil.copyfileobj(loader.byte_stream, target)
                loader.close()
                target.seek(0)
                loader = system.create_loader(File(source=target))
                if workbook_cache is not None:
                    workbook_cache[source] = target.name
                    atexit.register(os.remove, target.name)

        loader.open(mode='b')
        return loader

    def read_cell_stream_create(self):
        dialect = self.file.dialect

        # Get book
        # To fill merged cells we can't use read-only because
        # `sheet.merged_cell_ranges` is not available in this mode
        book = openpyxl.load_workbook(
            self.loader.byte_stream,
            read_only=not dialect.fill_merged_cells,
            data_only=True,
        )

        # Get sheet
        try:
            if isinstance(dialect.sheet, str):
                sheet = book[dialect.sheet]
            else:
                sheet = book.worksheets[dialect.sheet - 1]
        except (KeyError, IndexError):
            note = 'Excel document "%s" doesn\'t have a sheet "%s"'
            note = note % (self.file.source, dialect.sheet)
            raise exceptions.FrictionlessException(errors.SourceError(note=note))

        # Fill merged cells
        if dialect.fill_merged_cells:
            for merged_cell_range in list(sheet.merged_cells.ranges):
                merged_cell_range = str(merged_cell_range)
                sheet.unmerge_cells(merged_cell_range)
                merged_rows = openpyxl.utils.rows_from_range(merged_cell_range)
                coordinates = list(chain(*merged_rows))
                value = sheet[coordinates[0]].value
                for coordinate in coordinates:
                    cell = sheet[coordinate]
                    cell.value = value

        # Stream cells
        for row_number, row in enumerate(sheet.iter_rows(), start=1):
            yield (
                row_number,
                None,
                extract_row_values(
                    row, dialect.preserve_formatting, dialect.adjust_floating_point_error
                ),
            )

    # Write

    def write(self, source, target, headers, encoding=None):
        helpers.ensure_dir(target)
        count = 0
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=self.__options.get('sheet'))
        ws.append(headers)
        for row in source:
            ws.append(row)
            count += 1
        wb.save(target)
        return count


class XlsParser(Parser):
    Dialect = dialects.ExcelDialect
    loader_mode = 'b'

    # Read

    def read_cell_stream_create(self):
        dialect = self.file.dialect

        # Get book
        bytes = self.loader.byte_stream.read()
        try:
            book = xlrd.open_workbook(
                file_contents=bytes,
                encoding_override=self.file.encoding,
                formatting_info=True,
                logfile=sys.stderr,
            )
        except NotImplementedError:
            book = xlrd.open_workbook(
                file_contents=bytes,
                encoding_override=self.file.encoding,
                formatting_info=False,
                logfile=sys.stderr,
            )

        # Get sheet
        try:
            if isinstance(dialect.sheet, str):
                sheet = book.sheet_by_name(dialect.sheet)
            else:
                sheet = book.sheet_by_index(dialect.sheet)
        except (xlrd.XLRDError, IndexError):
            note = 'Excel document "%s" doesn\'t have a sheet "%s"'
            note = note % (self.file.source, dialect.sheet)
            raise exceptions.FrictionlessException(errors.SourceError(note=note))

        def type_value(ctype, value):
            """ Detects boolean value, int value, datetime """

            # Boolean
            if ctype == xlrd.XL_CELL_BOOLEAN:
                return bool(value)

            # Excel numbers are only float
            # Float with no decimals can be cast into int
            if ctype == xlrd.XL_CELL_NUMBER and value == value // 1:
                return int(value)

            # Datetime
            if ctype == xlrd.XL_CELL_DATE:
                return xlrd.xldate.xldate_as_datetime(value, self.__book.datemode)

            return value

        # Stream cells
        for x in range(0, sheet.nrows):
            row_number = x + 1
            row = []
            for y, value in enumerate(sheet.row_values(x)):
                value = type_value(sheet.cell(x, y).ctype, value)
                if dialect.fill_merged_cells:
                    for xlo, xhi, ylo, yhi in sheet.merged_cells:
                        if x in range(xlo, xhi) and y in range(ylo, yhi):
                            value = type_value(
                                sheet.cell(xlo, ylo).ctype, sheet.cell_value(xlo, ylo),
                            )
                row.append(value)
            yield (row_number, None, row)


# Internal

EXCEL_CODES = {
    "yyyy": "%Y",
    "yy": "%y",
    "dddd": "%A",
    "ddd": "%a",
    "dd": "%d",
    "d": "%-d",
    # Different from excel as there is no J-D in strftime
    "mmmmmm": "%b",
    "mmmm": "%B",
    "mmm": "%b",
    "hh": "%H",
    "h": "%-H",
    "ss": "%S",
    "s": "%-S",
    # Possibly different from excel as there is no am/pm in strftime
    "am/pm": "%p",
    # Different from excel as there is no A/P or a/p in strftime
    "a/p": "%p",
}

EXCEL_MINUTE_CODES = {
    "mm": "%M",
    "m": "%-M",
}
EXCEL_MONTH_CODES = {
    "mm": "%m",
    "m": "%-m",
}

EXCEL_MISC_CHARS = [
    "$",
    "+",
    "(",
    ":",
    "^",
    "'",
    "{",
    "<",
    "=",
    "-",
    "/",
    ")",
    "!",
    "&",
    "~",
    "}",
    ">",
    " ",
]

EXCEL_ESCAPE_CHAR = "\\"
EXCEL_SECTION_DIVIDER = ";"


def convert_excel_date_format_string(excel_date):
    """
    Created using documentation here:
    https://support.office.com/en-us/article/review-guidelines-for-customizing-a-number-format-c0a1d1fa-d3f4-4018-96b7-9c9354dd99f5

    """
    # The python date string that is being built
    python_date = ""
    # The excel code currently being parsed
    excel_code = ""
    prev_code = ""
    # If the previous character was the escape character
    char_escaped = False
    # If we are in a quotation block (surrounded by "")
    quotation_block = False
    # Variables used for checking if a code should be a minute or a month
    checking_minute_or_month = False
    minute_or_month_buffer = ""

    for c in excel_date:
        ec = excel_code.lower()
        # The previous character was an escape, the next character should be added normally
        if char_escaped:
            if checking_minute_or_month:
                minute_or_month_buffer += c
            else:
                python_date += c
            char_escaped = False
            continue
        # Inside a quotation block
        if quotation_block:
            if c == '"':
                # Quotation block should now end
                quotation_block = False
            elif checking_minute_or_month:
                minute_or_month_buffer += c
            else:
                python_date += c
            continue
        # The start of a quotation block
        if c == '"':
            quotation_block = True
            continue
        if c == EXCEL_SECTION_DIVIDER:
            # We ignore excel sections for datetimes
            break

        is_escape_char = c == EXCEL_ESCAPE_CHAR
        # The am/pm and a/p code add some complications, need to make sure we are not that code
        is_misc_char = c in EXCEL_MISC_CHARS and (c != "/" or (ec != "am" and ec != "a"))
        new_excel_code = False

        # Handle a new code without a different characeter in between
        if (
            ec
            and not is_escape_char
            and not is_misc_char
            # If the code does not start with c, we are in a new code
            and not ec.startswith(c.lower())
            # other than the case where we are building up
            # am/pm (minus the case where it is fully built), we are in a new code
            and (not ec.startswith("a") or ec == "am/pm")
        ):
            new_excel_code = True

        # Code is finished, check if it is a proper code
        if (is_escape_char or is_misc_char or new_excel_code) and ec:
            # Checking if the previous code should have been minute or month
            if checking_minute_or_month:
                if ec == "ss" or ec == "s":
                    # It should be a minute!
                    minute_or_month_buffer = (
                        EXCEL_MINUTE_CODES[prev_code] + minute_or_month_buffer
                    )
                else:
                    # It should be a months!
                    minute_or_month_buffer = (
                        EXCEL_MONTH_CODES[prev_code] + minute_or_month_buffer
                    )
                python_date += minute_or_month_buffer
                checking_minute_or_month = False
                minute_or_month_buffer = ""

            if ec in EXCEL_CODES:
                python_date += EXCEL_CODES[ec]
            # Handle months/minutes differently
            elif ec in EXCEL_MINUTE_CODES:
                # If preceded by hours, we know this is referring to minutes
                if prev_code == "h" or prev_code == "hh":
                    python_date += EXCEL_MINUTE_CODES[ec]
                else:
                    # Have to check if the next code is ss or s
                    checking_minute_or_month = True
                    minute_or_month_buffer = ""
            else:
                # Have to abandon this attempt to convert because the code is not recognized
                return None
            prev_code = ec
            excel_code = ""
        if is_escape_char:
            char_escaped = True
        elif is_misc_char:
            # Add the misc char
            if checking_minute_or_month:
                minute_or_month_buffer += c
            else:
                python_date += c
        else:
            # Just add to the code
            excel_code += c

    # Complete, check if there is still a buffer
    if checking_minute_or_month:
        # We know it's a month because there were no more codes after
        minute_or_month_buffer = EXCEL_MONTH_CODES[prev_code] + minute_or_month_buffer
        python_date += minute_or_month_buffer
    if excel_code:
        ec = excel_code.lower()
        if ec in EXCEL_CODES:
            python_date += EXCEL_CODES[ec]
        elif ec in EXCEL_MINUTE_CODES:
            if prev_code == "h" or prev_code == "hh":
                python_date += EXCEL_MINUTE_CODES[ec]
            else:
                python_date += EXCEL_MONTH_CODES[ec]
        else:
            return None
    return python_date


def convert_excel_number_format_string(excel_number, value):
    """
    A basic attempt to convert excel number_format to a number string

    The important goal here is to get proper amount of rounding
    """
    percentage = False
    if excel_number.endswith("%"):
        value = value * 100
        excel_number = excel_number[:-1]
        percentage = True
    if excel_number == "General":
        return value
    multi_codes = excel_number.split(";")
    if value < 0 and len(multi_codes) > 1:
        excel_number = multi_codes[1]
    else:
        excel_number = multi_codes[0]

    code = excel_number.split('.')

    if len(code) > 2:
        return None
    if len(code) < 2:
        # No decimals
        new_value = "{0:.0f}".format(value)
    else:
        decimal_section = code[1]
        # Only pay attention to the 0, # and ? characters as they provide precision information
        decimal_section = "".join(d for d in decimal_section if d in ["0", "#", "?"])

        # Count the number of hashes at the end of the decimal_section in order to know how
        # the number should be truncated
        number_hash = 0
        for i in reversed(range(len(decimal_section))):
            if decimal_section[i] == "#":
                number_hash += 1
            else:
                break
        string_format_code = "{0:." + str(len(decimal_section)) + "f}"
        new_value = string_format_code.format(value)
        if number_hash > 0:
            for i in range(number_hash):
                if new_value.endswith("0"):
                    new_value = new_value[:-1]
    if percentage:
        return new_value + "%"

    return new_value


def extract_row_values(row, preserve_formatting=False, adjust_floating_point_error=False):
    if preserve_formatting:
        values = []
        for cell in row:
            number_format = cell.number_format or ""
            value = cell.value

            if isinstance(cell.value, datetime.datetime) or isinstance(
                cell.value, datetime.time
            ):
                temporal_format = convert_excel_date_format_string(number_format)
                if temporal_format:
                    value = cell.value.strftime(temporal_format)
            elif (
                adjust_floating_point_error
                and isinstance(cell.value, float)
                and number_format == "General"
            ):
                # We have a float with format General
                # Calculate the number of integer digits
                integer_digits = len(str(int(cell.value)))
                # Set the precision to 15 minus the number of integer digits
                precision = 15 - (integer_digits)
                value = round(cell.value, precision)
            elif isinstance(cell.value, (int, float)):
                new_value = convert_excel_number_format_string(number_format, cell.value)
                if new_value:
                    value = new_value
            values.append(value)
        return values
    return list(cell.value for cell in row)